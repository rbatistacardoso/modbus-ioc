from openpyxl import load_workbook
from string import Template

ai_template = Template(
    '''
    record(ai, "${pv_name}") {
        field(DESC, "")
        field(EGU, "${unit}")
        field(DTYP,"asynInt32")
        field(INP,"@asyn($(PORT), ${register_addr})${data_type}")
        field(ASLO, "${resolution}")
        field(SCAN,"2 second")
    }
    '''
)

SHEET = load_workbook(filename='../etc/M1M Modbus map V1.2.xlsx')['M1M registers map']

## Create a dictionary of column names
ColNames = {}
Current  = 0
for COL in SHEET.iter_cols(1, SHEET.max_column):
    ColNames[COL[0].value] = Current
    Current += 1

# Temporary
def to_camel(string: str):
    if type(string) != None:
        output = ''
        for _ in range(len(string)):
            if _>0 and string[_-1]==' ':
                output += string[_].upper()
            else:
                output += string[_]
        output = output.replace(" ", "")
        return output

with open('../ioc/database/st.db', 'w') as file:

    for row_cells in SHEET.iter_rows(2, SHEET.max_row):
        
        if row_cells[ColNames['Monitored']].value == 'yes':
            
            if row_cells[ColNames['Quantity/Functionality']].value != None:
            
                if row_cells[ColNames['Data Type']].value == 'Unsigned' and row_cells[ColNames['Length']].value == 2: dt = ''
                elif row_cells[ColNames['Data Type']].value == 'Unsigned' and row_cells[ColNames['Length']].value == 1: dt = 'INT16'
                elif row_cells[ColNames['Data Type']].value == 'Signed' and row_cells[ColNames['Length']].value == 2: dt = 'INT32_BE'
                elif row_cells[ColNames['Data Type']].value == 'Signed' and row_cells[ColNames['Length']].value == 1: dt = 'INT32_BE'
            
                file.write(ai_template.safe_substitute(
                    pv_name = to_camel(row_cells[ColNames['Quantity/Functionality']].value),
                    register_addr = row_cells[ColNames['Register\n(Dec)']].value,
                    data_type = '',
                    resolution = row_cells[ColNames['Resolution']].value,
                    unit = row_cells[ColNames['Unit']].value
                    ))
